"""
Excel Export Service — generates .xlsx files from analysis data.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


SEVERITY_FILLS = {
    3: PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid'),  # Red
    2: PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),  # Yellow
    1: PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid'),  # Blue
}

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT_WHITE = Font(bold=True, color='FFFFFF')


def _write_headers(ws, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')


def export_bpa(violations):
    wb = Workbook()
    ws = wb.active
    ws.title = 'BPA Violations'

    headers = ['Severity', 'Category', 'Rule Name', 'Object', 'Object Type', 'Table', 'Description', 'Fix Action', 'Fix Steps']
    _write_headers(ws, headers)

    for i, v in enumerate(violations, 2):
        ws.cell(row=i, column=1, value=v.severity_label)
        ws.cell(row=i, column=2, value=v.category)
        ws.cell(row=i, column=3, value=v.rule_name)
        ws.cell(row=i, column=4, value=v.object_name)
        ws.cell(row=i, column=5, value=v.object_type)
        ws.cell(row=i, column=6, value=v.table_name)
        ws.cell(row=i, column=7, value=v.description)

        fix = v.fix_steps or {}
        ws.cell(row=i, column=8, value=fix.get('action', '') if isinstance(fix, dict) else '')
        steps = fix.get('steps', []) if isinstance(fix, dict) else []
        ws.cell(row=i, column=9, value='\n'.join(steps) if steps else '')

        fill = SEVERITY_FILLS.get(v.severity)
        if fill:
            ws.cell(row=i, column=1).fill = fill

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_inspector(results):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inspector Results'

    headers = ['Status', 'Rule Name', 'Description', 'Page', 'Expected', 'Actual']
    _write_headers(ws, headers)

    for i, r in enumerate(results, 2):
        ws.cell(row=i, column=1, value='PASS' if r.passed else 'FAIL')
        ws.cell(row=i, column=2, value=r.rule_name)
        ws.cell(row=i, column=3, value=r.rule_description)
        ws.cell(row=i, column=4, value=r.page_name)
        ws.cell(row=i, column=5, value=str(r.expected) if r.expected else '')
        ws.cell(row=i, column=6, value=str(r.actual) if r.actual else '')

        if not r.passed:
            ws.cell(row=i, column=1).fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_catalog(tables, columns, measures, relationships, roles):
    wb = Workbook()

    # Tables sheet
    ws = wb.active
    ws.title = 'Tables'
    _write_headers(ws, ['Table Name', 'Mode', 'Hidden', 'Columns', 'Measures', 'Partitions'])
    for i, t in enumerate(tables, 2):
        ws.cell(row=i, column=1, value=t.name)
        ws.cell(row=i, column=2, value=t.mode)
        ws.cell(row=i, column=3, value='Yes' if t.is_hidden else 'No')
        ws.cell(row=i, column=4, value=t.column_count)
        ws.cell(row=i, column=5, value=t.measure_count)
        ws.cell(row=i, column=6, value=t.partition_count)

    # Columns sheet
    ws2 = wb.create_sheet('Columns')
    _write_headers(ws2, ['Table', 'Column Name', 'Data Type', 'Hidden', 'Format String', 'Display Folder', 'Source Column'])
    row = 2
    for c in columns:
        table_name = c.table.name if c.table else ''
        ws2.cell(row=row, column=1, value=table_name)
        ws2.cell(row=row, column=2, value=c.name)
        ws2.cell(row=row, column=3, value=c.data_type)
        ws2.cell(row=row, column=4, value='Yes' if c.is_hidden else 'No')
        ws2.cell(row=row, column=5, value=c.format_string)
        ws2.cell(row=row, column=6, value=c.display_folder)
        ws2.cell(row=row, column=7, value=c.source_column)
        row += 1

    # Measures sheet
    ws3 = wb.create_sheet('Measures')
    _write_headers(ws3, ['Table', 'Measure Name', 'Expression', 'Format String', 'Display Folder', 'Hidden'])
    row = 2
    for m in measures:
        table_name = m.table.name if m.table else ''
        ws3.cell(row=row, column=1, value=table_name)
        ws3.cell(row=row, column=2, value=m.name)
        ws3.cell(row=row, column=3, value=m.expression)
        ws3.cell(row=row, column=4, value=m.format_string)
        ws3.cell(row=row, column=5, value=m.display_folder)
        ws3.cell(row=row, column=6, value='Yes' if m.is_hidden else 'No')
        row += 1

    # Relationships sheet
    ws4 = wb.create_sheet('Relationships')
    _write_headers(ws4, ['From Table', 'From Column', 'To Table', 'To Column', 'Cross Filter', 'From Cardinality', 'To Cardinality', 'Active'])
    for i, r in enumerate(relationships, 2):
        ws4.cell(row=i, column=1, value=r.from_table)
        ws4.cell(row=i, column=2, value=r.from_column)
        ws4.cell(row=i, column=3, value=r.to_table)
        ws4.cell(row=i, column=4, value=r.to_column)
        ws4.cell(row=i, column=5, value=r.cross_filter)
        ws4.cell(row=i, column=6, value=r.from_cardinality)
        ws4.cell(row=i, column=7, value=r.to_cardinality)
        ws4.cell(row=i, column=8, value='Yes' if r.is_active else 'No')

    # Roles sheet
    ws5 = wb.create_sheet('Roles')
    _write_headers(ws5, ['Role Name', 'Table', 'Filter Expression'])
    row = 2
    for role in roles:
        filters = role.filters or []
        for f in filters:
            ws5.cell(row=row, column=1, value=role.name)
            ws5.cell(row=row, column=2, value=f.get('table', ''))
            ws5.cell(row=row, column=3, value=f.get('expression', ''))
            row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
