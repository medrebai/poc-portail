"""
bpa-to-excel.py
Converts bpa-results.json into a formatted bpa-report.xlsx.
Standalone -- works with just the BPA results file.

Usage:
    python ci-tools/bpa-to-excel.py
    python ci-tools/bpa-to-excel.py --input ci-tools/bpa-results-new.json --output ci-tools/bpa-report.xlsx
"""

import json
import argparse
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY     = "1E3A5F"
WHITE    = "FFFFFF"
LGRAY    = "F7F9FC"
WARN_BG  = "FEF9E7"
WARN_FG  = "B7770D"
ERR_BG   = "FDECEA"
ERR_FG   = "C0392B"
OK_BG    = "EAFAF1"
OK_FG    = "1E8449"
INFO_BG  = "EBF3FB"
INFO_FG  = "2E5FA3"
TITLE_FG = "1E3A5F"

THIN   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SEV_COLORS = {
    "ERROR":   (ERR_BG, ERR_FG),
    "WARNING": (WARN_BG, WARN_FG),
    "INFO":    (INFO_BG, INFO_FG),
    "UNKNOWN": (LGRAY, "666666"),
}


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def hcell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=True, color=WHITE, name="Arial", size=10)
    c.fill      = fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def dcell(ws, row, col, value, bold=False, fg=None, bg=None, align="left", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, color=fg or "000000", name="Arial", size=10)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border    = BORDER
    if bg:
        c.fill = fill(bg)
    return c


def alt(idx):
    return LGRAY if idx % 2 == 0 else WHITE


def sheet_title(ws, title, subtitle=""):
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = title
    c.font      = Font(bold=True, size=14, color=TITLE_FG, name="Arial")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    if subtitle:
        ws.merge_cells("A2:H2")
        s = ws["A2"]
        s.value     = subtitle
        s.font      = Font(size=9, color="888888", name="Arial", italic=True)
        s.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 16
        return 4
    return 3


def build_summary_sheet(wb, data):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["D"].width = 18

    sev = data.get("summaryBySeverity", [])
    total = data.get("totalViolations", 0)
    model = data.get("model", "N/A")
    rules_file = data.get("rulesFile", "N/A")
    ts = data.get("timestamp", "N/A")

    start = sheet_title(ws, "BPA Quality Report",
        "Model: {}  |  Rules: {}  |  {}".format(model, rules_file, ts))

    ws.merge_cells("B{}:D{}".format(start, start))
    c = ws.cell(row=start, column=2, value="TOTAL VIOLATIONS: {}".format(total))
    c.font = Font(bold=True, size=12, color=WHITE, name="Arial")
    c.fill = fill(ERR_FG if total > 0 else OK_FG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start].height = 28
    start += 2

    ws.merge_cells("B{}:D{}".format(start, start))
    c = ws.cell(row=start, column=2, value="BREAKDOWN BY SEVERITY")
    c.font = Font(bold=True, size=11, color=TITLE_FG, name="Arial")
    c.fill = fill("D6E4F0")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[start].height = 22
    start += 1

    for s in sorted(sev, key=lambda x: x.get("severity", 0), reverse=True):
        label = s.get("label", "UNKNOWN")
        count = s.get("count", 0)
        sbg, sfg = SEV_COLORS.get(label, (LGRAY, "000000"))
        dcell(ws, start, 2, label, bold=True, fg=sfg, bg=sbg, align="left")
        dcell(ws, start, 4, count, bold=True, fg=sfg, bg=sbg, align="center")
        ws.row_dimensions[start].height = 20
        start += 1

    start += 1
    by_rule = data.get("summaryByRule", [])
    ws.merge_cells("B{}:D{}".format(start, start))
    c = ws.cell(row=start, column=2, value="RULES VIOLATED: {}".format(len(by_rule)))
    c.font = Font(bold=True, size=11, color=TITLE_FG, name="Arial")
    c.fill = fill("D6E4F0")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[start].height = 22


def build_violations_sheet(wb, data):
    ws = wb.create_sheet("Violations")
    ws.sheet_view.showGridLines = False

    violations = data.get("violations", [])
    total = data.get("totalViolations", len(violations))
    model = data.get("model", "N/A")

    start = sheet_title(ws, "BPA Violations Detail",
        "{} violations  |  Model: {}".format(total, model))

    headers = ["Severity", "Category", "Rule", "Object", "Type", "Description", "Suggested Fix"]
    widths  = [14, 18, 42, 38, 14, 50, 55]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, start, i, h, w)
    ws.row_dimensions[start].height = 20
    ws.freeze_panes = "A{}".format(start + 1)

    for idx, v in enumerate(violations, 1):
        row = start + idx
        rbg = alt(idx)
        sev = v.get("severityLabel", "UNKNOWN")
        sbg, sfg = SEV_COLORS.get(sev, (rbg, "000000"))

        dcell(ws, row, 1, sev, bold=True, fg=sfg, bg=sbg, align="center")
        dcell(ws, row, 2, v.get("category", ""), bg=rbg)
        dcell(ws, row, 3, v.get("ruleName", ""), bg=rbg)
        dcell(ws, row, 4, v.get("object", ""), bg=rbg)
        dcell(ws, row, 5, v.get("objectType", ""), bg=rbg, align="center")
        dcell(ws, row, 6, v.get("description", ""), bg=rbg, wrap=True)

        fix = v.get("suggestedFix", {})
        fix_text = ""
        if isinstance(fix, dict):
            action = fix.get("action", "")
            steps = fix.get("steps", [])
            if steps:
                numbered = "  ".join("{}. {}".format(i+1, s) for i, s in enumerate(steps))
                if action and action != "Manual review required":
                    fix_text = action + " -- " + numbered
                else:
                    fix_text = numbered
            elif action:
                fix_text = action
        elif fix:
            fix_text = str(fix)

        dcell(ws, row, 7, fix_text, bg=rbg, wrap=True)
        ws.row_dimensions[row].height = 45 if (isinstance(fix, dict) and fix.get("steps")) else 30


def build_by_rule_sheet(wb, data):
    ws = wb.create_sheet("By Rule")
    ws.sheet_view.showGridLines = False

    by_rule = data.get("summaryByRule", [])

    start = sheet_title(ws, "BPA Summary by Rule",
        "{} rules violated".format(len(by_rule)))

    headers = ["Count", "Severity", "Category", "Rule", "Rule ID"]
    widths  = [10, 14, 18, 52, 48]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, start, i, h, w)
    ws.row_dimensions[start].height = 20
    ws.freeze_panes = "A{}".format(start + 1)

    sorted_rules = sorted(by_rule, key=lambda x: (-x.get("severity", 0), -x.get("count", 0)))

    for idx, rule in enumerate(sorted_rules, 1):
        row = start + idx
        rbg = alt(idx)
        sev = rule.get("severityLabel", "UNKNOWN")
        sbg, sfg = SEV_COLORS.get(sev, (rbg, "000000"))

        dcell(ws, row, 1, rule.get("count", 0), bold=True, bg=rbg, align="center")
        dcell(ws, row, 2, sev, bold=True, fg=sfg, bg=sbg, align="center")
        dcell(ws, row, 3, rule.get("category", ""), bg=rbg)
        dcell(ws, row, 4, rule.get("ruleName", ""), bg=rbg)
        dcell(ws, row, 5, rule.get("ruleId", ""), bg=rbg)
        ws.row_dimensions[row].height = 18


def main():
    parser = argparse.ArgumentParser(description="Convert bpa-results.json to Excel.")
    parser.add_argument("--input",  default="ci-tools/bpa-results.json")
    parser.add_argument("--output", default="ci-tools/bpa-report.xlsx")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print("ERROR: BPA results not found: {}".format(input_path))
        sys.exit(1)

    print("Reading: {}".format(input_path))
    data = json.loads(input_path.read_text(encoding="utf-8-sig"))

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print("Building sheets...")
    build_summary_sheet(wb, data)
    build_violations_sheet(wb, data)
    build_by_rule_sheet(wb, data)

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))

    print("Excel report saved to: {}".format(out_path))
    print("Sheets: {}".format(", ".join(wb.sheetnames)))


if __name__ == "__main__":
    main()