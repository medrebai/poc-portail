"""
json-to-excel.py
Converts quality-summary.json into a formatted quality-report.xlsx.
Run locally — not part of the CI pipeline.

Usage:
    python ci-tools/json-to-excel.py
    python ci-tools/json-to-excel.py \
        --summary ci-tools/quality-summary.json \
        --output  ci-tools/quality-report.xlsx
"""

import json
import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Style constants ───────────────────────────────────────────────────────
NAVY     = "1E3A5F"
WHITE    = "FFFFFF"
LGRAY    = "F7F9FC"
WARN_BG  = "FEF9E7"
WARN_FG  = "B7770D"
ERR_BG   = "FDECEA"
ERR_FG   = "C0392B"
OK_BG    = "EAFAF1"
OK_FG    = "1E8449"
INFO_BG  = "EBF3FB"
INFO_FG  = "2E5FA3"
TITLE_FG = "1E3A5F"

THIN   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def hcell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=True, color=WHITE, name="Arial", size=10)
    c.fill      = fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def dcell(ws, row, col, value, bold=False, fg=None, bg=None, align="left", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, color=fg or "000000", name="Arial", size=10)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border    = BORDER
    if bg:
        c.fill = fill(bg)
    return c


def alt(idx):
    return LGRAY if idx % 2 == 0 else WHITE


def sheet_title(ws, title, subtitle=""):
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = title
    c.font      = Font(bold=True, size=14, color=TITLE_FG, name="Arial")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    if subtitle:
        ws.merge_cells("A2:H2")
        s = ws["A2"]
        s.value     = subtitle
        s.font      = Font(size=9, color="888888", name="Arial", italic=True)
        s.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 16
        return 3
    return 2


# ═══════════════════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ═══════════════════════════════════════════════════════════════════════════

def build_summary_sheet(wb, data):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["D"].width = 18

    summary   = data.get("summary", {})
    bpa       = summary.get("bpa", {})
    inspector = summary.get("inspector", {})

    start = sheet_title(ws, "Quality Report Summary",
        f"Branch: {data.get('branch', 'N/A')}  |  "
        f"PR: {data.get('pullRequestId', 'N/A')}  |  "
        f"Duration: {data.get('durationSeconds', 0)}s  |  "
        f"Run: {data.get('pipelineRunId', 'local')}")

    sections = [
        ("BPA - SEMANTIC MODEL", None),
        ("Total violations",           bpa.get("totalViolations", 0)),
        ("Errors",                     bpa.get("errors", 0)),
        ("Warnings",                   bpa.get("warnings", 0)),
        ("", None),
        ("PBI INSPECTOR - VISUAL LAYER", None),
        ("Total rules checked",        inspector.get("totalRules", 0)),
        ("Rules failed",               inspector.get("failed", 0)),
        ("Rules passed",               inspector.get("passed", 0)),
    ]

    row = start
    for label, value in sections:
        if value is None and label:
            ws.merge_cells(f"B{row}:D{row}")
            c = ws.cell(row=row, column=2, value=label)
            c.font = Font(bold=True, size=11, color=TITLE_FG, name="Arial")
            c.fill = fill("D6E4F0")
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 22
        elif label == "":
            ws.row_dimensions[row].height = 8
        else:
            lc = ws.cell(row=row, column=2, value=label)
            lc.font   = Font(name="Arial", size=10)
            lc.border = BORDER
            lc.alignment = Alignment(horizontal="left", vertical="center")

            # Color coding
            vbg, vfg = WHITE, "000000"
            if isinstance(value, int) and value > 0:
                if "Error" in label or "failed" in label.lower() or "violations" in label.lower():
                    vbg, vfg = WARN_BG, WARN_FG

            vc = ws.cell(row=row, column=4, value=value)
            vc.font = Font(bold=True, name="Arial", size=10, color=vfg)
            vc.fill = fill(vbg)
            vc.alignment = Alignment(horizontal="center", vertical="center")
            vc.border = BORDER
            ws.row_dimensions[row].height = 20
        row += 1


def build_bpa_violations_sheet(wb, data):
    ws = wb.create_sheet("BPA Violations")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    model_analysis = data.get("modelAnalysis", {})
    violations     = model_analysis.get("violations", [])
    total          = model_analysis.get("totalViolations", len(violations))

    start = sheet_title(ws, "BPA Violations",
        f"{total} violations  |  Model: {model_analysis.get('model', 'N/A')}")

    headers = ["Severity", "Category", "Rule", "Object", "Type", "Description", "Suggested Fix"]
    widths  = [12, 22, 38, 36, 14, 40, 55]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, start, i, h, w)
    ws.row_dimensions[start].height = 20

    sev_color = {
        "ERROR":   (ERR_BG,  ERR_FG),
        "WARNING": (WARN_BG, WARN_FG),
        "INFO":    (INFO_BG, INFO_FG),
    }

    for idx, v in enumerate(violations, 1):
        row = start + idx
        rbg = alt(idx)
        sev = v.get("severityLabel", "INFO")
        sbg, sfg = sev_color.get(sev, (rbg, "000000"))

        dcell(ws, row, 1, sev, bold=True, fg=sfg, bg=sbg, align="center")
        dcell(ws, row, 2, v.get("category", ""), bg=rbg)
        dcell(ws, row, 3, v.get("ruleName", ""), bg=rbg)
        dcell(ws, row, 4, v.get("object", ""), bg=rbg)
        dcell(ws, row, 5, v.get("objectType", ""), bg=rbg, align="center")
        dcell(ws, row, 6, v.get("description", ""), bg=rbg, wrap=True)

        fix = v.get("suggestedFix", {})
        fix_text = fix.get("action", "")
        if fix.get("steps"):
            fix_text += "\n" + "\n".join(f"  {i+1}. {s}" for i, s in enumerate(fix["steps"]))
        dcell(ws, row, 7, fix_text, bg=rbg, wrap=True)
        ws.row_dimensions[row].height = 45 if fix.get("steps") else 17


def build_bpa_summary_sheet(wb, data):
    ws = wb.create_sheet("BPA By Rule")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    model_analysis = data.get("modelAnalysis", {})
    by_rule        = model_analysis.get("summaryByRule", [])

    start = sheet_title(ws, "BPA Summary by Rule",
        f"{len(by_rule)} rules violated")

    headers = ["Count", "Severity", "Category", "Rule", "Rule ID"]
    widths  = [10, 14, 22, 50, 42]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, start, i, h, w)
    ws.row_dimensions[start].height = 20

    sev_color = {
        "ERROR":   (ERR_BG,  ERR_FG),
        "WARNING": (WARN_BG, WARN_FG),
        "INFO":    (INFO_BG, INFO_FG),
    }

    for idx, r in enumerate(by_rule, 1):
        row = start + idx
        rbg = alt(idx)
        sev = r.get("severityLabel", "INFO")
        sbg, sfg = sev_color.get(sev, (rbg, "000000"))

        dcell(ws, row, 1, r.get("count", 0), bold=True, bg=rbg, align="center")
        dcell(ws, row, 2, sev, bold=True, fg=sfg, bg=sbg, align="center")
        dcell(ws, row, 3, r.get("category", ""), bg=rbg)
        dcell(ws, row, 4, r.get("ruleName", ""), bg=rbg)
        dcell(ws, row, 5, r.get("ruleId", ""), bg=rbg)
        ws.row_dimensions[row].height = 17


def build_inspector_sheet(wb, data):
    ws = wb.create_sheet("PBI Inspector")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    visual_analysis = data.get("visualAnalysis", {})
    results         = visual_analysis.get("Results", [])

    failed = [r for r in results if not r.get("Pass", True)]
    passed = [r for r in results if r.get("Pass", True)]

    start = sheet_title(ws, "PBI Inspector Results",
        f"{len(results)} rules  |  {len(passed)} passed  |  {len(failed)} failed")

    headers = ["Status", "Rule ID", "Rule Name", "Page", "Description", "Actual"]
    widths  = [12, 36, 40, 28, 45, 45]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, start, i, h, w)
    ws.row_dimensions[start].height = 20

    # Show failed first
    all_results = sorted(results, key=lambda r: (1 if r.get("Pass") else 0))

    for idx, r in enumerate(all_results, 1):
        row   = start + idx
        rbg   = alt(idx)
        passed = r.get("Pass", False)

        if passed:
            dcell(ws, row, 1, "PASSED", bold=True, fg=OK_FG, bg=OK_BG, align="center")
        else:
            dcell(ws, row, 1, "FAILED", bold=True, fg=ERR_FG, bg=ERR_BG, align="center")

        dcell(ws, row, 2, r.get("RuleId", ""), bg=rbg)
        dcell(ws, row, 3, r.get("RuleName", ""), bg=rbg)
        dcell(ws, row, 4, r.get("ParentDisplayName", "N/A"), bg=rbg)
        dcell(ws, row, 5, r.get("RuleDescription", ""), bg=rbg, wrap=True)

        actual = r.get("Actual", "")
        if isinstance(actual, list):
            actual = "\n".join(str(a) for a in actual)
        elif isinstance(actual, dict):
            actual = json.dumps(actual, indent=2)
        dcell(ws, row, 6, str(actual) if actual else "", bg=rbg, wrap=True)
        ws.row_dimensions[row].height = 35 if not passed else 17


def build_model_overview_sheet(wb, data):
    ws = wb.create_sheet("Model Overview")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["D"].width = 16

    catalog = data.get("modelCatalog", {})
    model   = catalog.get("model", {})
    tables  = catalog.get("tables", [])
    rels    = catalog.get("relationships", [])
    roles   = catalog.get("roles", [])

    all_cols  = [c for t in tables for c in t.get("columns", [])]
    all_meas  = [m for t in tables for m in t.get("measures", [])]
    all_parts = [p for t in tables for p in t.get("partitions", [])]
    bq_parts  = [p for p in all_parts if p.get("datasource") == "Google BigQuery"]
    float_c   = [c for c in all_cols if c.get("isFloat")]
    miss_fmt  = [c for c in all_cols if c.get("missingFormatString")]
    bidir     = [r for r in rels if r.get("crossFilter") == "bothDirections"]

    start = sheet_title(ws, "Model Overview",
        f"Culture: {model.get('culture', 'N/A')}  |  "
        f"Auto Date/Time: {'YES' if model.get('autoDateTime') else 'NO'}")

    sections = [
        ("MODEL STRUCTURE", None),
        ("Tables",          len(tables)),
        ("Columns",         len(all_cols)),
        ("Measures",        len(all_meas)),
        ("Relationships",   len(rels)),
        ("Roles RLS",       len(roles)),
        ("BigQuery sources", len(bq_parts)),
        ("", None),
        ("QUALITY FLAGS", None),
        ("Auto Date/Time enabled",       "YES" if model.get("autoDateTime") else "NO"),
        ("Float (Double) columns",       len(float_c)),
        ("Columns missing format string", len(miss_fmt)),
        ("Bidirectional relationships",  len(bidir)),
    ]

    row = start
    for label, value in sections:
        if value is None and label:
            ws.merge_cells(f"B{row}:D{row}")
            c = ws.cell(row=row, column=2, value=label)
            c.font = Font(bold=True, size=11, color=TITLE_FG, name="Arial")
            c.fill = fill("D6E4F0")
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 22
        elif label == "":
            ws.row_dimensions[row].height = 8
        else:
            lc = ws.cell(row=row, column=2, value=label)
            lc.font   = Font(name="Arial", size=10)
            lc.border = BORDER
            lc.alignment = Alignment(horizontal="left", vertical="center")

            vbg, vfg = WHITE, "000000"
            if isinstance(value, int) and value > 0:
                if any(k in label for k in ["Float", "format string", "Bidirectional"]):
                    vbg, vfg = WARN_BG, WARN_FG
            if value == "YES" and "Auto Date" in label:
                vbg, vfg = WARN_BG, WARN_FG
            if value == "NO" and "Auto Date" in label:
                vbg, vfg = OK_BG, OK_FG

            vc = ws.cell(row=row, column=4, value=value)
            vc.font = Font(bold=True, name="Arial", size=10, color=vfg)
            vc.fill = fill(vbg)
            vc.alignment = Alignment(horizontal="center", vertical="center")
            vc.border = BORDER
            ws.row_dimensions[row].height = 20
        row += 1


def build_columns_sheet(wb, data):
    ws = wb.create_sheet("Tables & Columns")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    catalog = data.get("modelCatalog", {})
    tables  = catalog.get("tables", [])
    total   = sum(len(t.get("columns", [])) for t in tables)

    start = sheet_title(ws, "Tables & Columns",
        f"{len(tables)} tables  |  {total} columns total")

    headers = ["Table", "Column", "Data Type", "Hidden", "Format String",
               "Calculated", "Float", "Missing Format"]
    widths  = [28, 34, 16, 10, 18, 12, 10, 18]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, start, i, h, w)
    ws.row_dimensions[start].height = 20

    row = start + 1
    for t in tables:
        for c in t.get("columns", []):
            rbg = alt(row)
            dcell(ws, row, 1, t["name"], bg=rbg)
            dcell(ws, row, 2, c["name"], bg=rbg)
            dcell(ws, row, 3, c.get("dataType", ""), bg=rbg, align="center")
            dcell(ws, row, 4, "Yes" if c.get("isHidden") else "No", bg=rbg, align="center")
            dcell(ws, row, 5, c.get("formatString") or "-", bg=rbg)
            dcell(ws, row, 6, "Yes" if c.get("isCalculated") else "No", bg=rbg, align="center")

            if c.get("isFloat"):
                dcell(ws, row, 7, "FLOAT", bold=True, fg=WARN_FG, bg=WARN_BG, align="center")
            else:
                dcell(ws, row, 7, "", bg=rbg)

            if c.get("missingFormatString"):
                dcell(ws, row, 8, "MISSING", bold=True, fg=WARN_FG, bg=WARN_BG, align="center")
            else:
                dcell(ws, row, 8, "", bg=rbg)

            ws.row_dimensions[row].height = 17
            row += 1


def build_measures_sheet(wb, data):
    ws = wb.create_sheet("Measures")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    catalog  = data.get("modelCatalog", {})
    tables   = catalog.get("tables", [])
    all_meas = [(t["name"], m) for t in tables for m in t.get("measures", [])]

    start = sheet_title(ws, "Measures DAX",
        f"{len(all_meas)} measures total")

    headers = ["Table", "Measure", "Type", "Format String", "Folder", "Hidden", "Expression DAX", "Missing Format"]
    widths  = [24, 32, 14, 18, 22, 10, 55, 16]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, start, i, h, w)
    ws.row_dimensions[start].height = 20

    for idx, (tname, m) in enumerate(all_meas, 1):
        row = start + idx
        rbg = alt(idx)
        dcell(ws, row, 1, tname, bg=rbg)
        dcell(ws, row, 2, m["name"], bg=rbg)
        dcell(ws, row, 3, m.get("dataType") or "-", bg=rbg, align="center")
        dcell(ws, row, 4, m.get("formatString") or "-", bg=rbg)
        dcell(ws, row, 5, m.get("displayFolder") or "-", bg=rbg)
        dcell(ws, row, 6, "Yes" if m.get("isHidden") else "No", bg=rbg, align="center")
        dcell(ws, row, 7, m.get("expression", ""), bg=rbg, wrap=True)
        if m.get("missingFormatString"):
            dcell(ws, row, 8, "MISSING", bold=True, fg=WARN_FG, bg=WARN_BG, align="center")
        else:
            dcell(ws, row, 8, "", bg=rbg)
        ws.row_dimensions[row].height = 30


def build_relationships_sheet(wb, data):
    ws = wb.create_sheet("Relationships")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    catalog = data.get("modelCatalog", {})
    rels    = catalog.get("relationships", [])

    start = sheet_title(ws, "Relationships",
        f"{len(rels)} relationships")

    headers = ["From Table", "From Column", "Cardinality", "To Table", "To Column", "Cross Filter", "Active"]
    widths  = [28, 30, 16, 28, 30, 20, 10]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, start, i, h, w)
    ws.row_dimensions[start].height = 20

    for idx, r in enumerate(rels, 1):
        row = start + idx
        rbg = alt(idx)
        card = f"{r.get('fromCardinality', 'many')}:{r.get('toCardinality', 'one')}"
        dcell(ws, row, 1, r.get("fromTable", ""), bg=rbg)
        dcell(ws, row, 2, r.get("fromColumn", ""), bg=rbg)
        dcell(ws, row, 3, card, bg=rbg, align="center")
        dcell(ws, row, 4, r.get("toTable", ""), bg=rbg)
        dcell(ws, row, 5, r.get("toColumn", ""), bg=rbg)
        if r.get("crossFilter") == "bothDirections":
            dcell(ws, row, 6, "Bidirectional", bold=True, fg=WARN_FG, bg=WARN_BG, align="center")
        else:
            dcell(ws, row, 6, "Single", bg=rbg, align="center")
        if r.get("isActive"):
            dcell(ws, row, 7, "Yes", fg=OK_FG, bg=OK_BG, align="center")
        else:
            dcell(ws, row, 7, "No", fg=ERR_FG, bg=ERR_BG, align="center")
        ws.row_dimensions[row].height = 17


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def build_sources_sheet(wb, data):
    ws = wb.create_sheet("Sources & Partitions")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    catalog   = data.get("modelCatalog", {})
    tables    = catalog.get("tables", [])
    all_parts = [(t["name"], p) for t in tables for p in t.get("partitions", [])]

    bq_count   = sum(1 for _, p in all_parts if p.get("datasource") == "Google BigQuery")
    calc_count = sum(1 for _, p in all_parts if p.get("queryType") == "calculated")

    start = sheet_title(ws, "Data Sources & M Queries",
        f"{len(all_parts)} partition(s)  |  {bq_count} BigQuery  |  {calc_count} calculated tables")

    headers = ["Table", "Source", "BQ Project", "BQ Dataset", "BQ Table",
               "SQL Query", "M Transformations", "Steps", "M Query"]
    widths  = [26, 18, 18, 20, 24, 45, 38, 8, 70]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, start, i, h, w)
    ws.row_dimensions[start].height = 20

    for idx, (tname, p) in enumerate(all_parts, 1):
        row = start + idx
        rbg = alt(idx)
        ds = p.get("datasource", "Unknown")
        if ds == "Google BigQuery":
            src_fg, src_bg = INFO_FG, INFO_BG
        elif p.get("queryType") == "calculated":
            src_fg, src_bg = OK_FG, OK_BG
        else:
            src_fg, src_bg = "000000", rbg

        dcell(ws, row, 1, tname, bg=rbg)
        dcell(ws, row, 2, ds, bold=True, fg=src_fg, bg=src_bg, align="center")
        dcell(ws, row, 3, p.get("bqProject") or "-", bg=rbg)
        dcell(ws, row, 4, p.get("bqDataset") or "-", bg=rbg)
        dcell(ws, row, 5, p.get("bqTable")   or "-", bg=rbg)
        dcell(ws, row, 6, p.get("sqlQuery")  or "-", bg=rbg, wrap=True)
        transforms = p.get("transformations", [])
        dcell(ws, row, 7, ", ".join(transforms) if transforms else "-", bg=rbg, wrap=True)
        dcell(ws, row, 8, p.get("stepCount") or "-", bg=rbg, align="center")
        dcell(ws, row, 9, p.get("mQuery") or "-", bg=rbg, wrap=True)
        ws.row_dimensions[row].height = 55 if p.get("mQuery") else 17



def build_roles_sheet(wb, data):
    ws = wb.create_sheet("Roles RLS")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    catalog = data.get("modelCatalog", {})
    roles   = catalog.get("roles", [])

    start = sheet_title(ws, "Security Roles (RLS)",
        f"{len(roles)} role(s) defined")

    headers = ["Role", "Filtered Table", "DAX Expression"]
    widths  = [28, 32, 60]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, start, i, h, w)
    ws.row_dimensions[start].height = 20

    if not roles:
        c = ws.cell(row=start + 1, column=1, value="No RLS roles defined in this model.")
        c.font = Font(italic=True, color="888888", name="Arial", size=10)
        return

    row = start + 1
    for role in roles:
        filters = role.get("filters", [])
        if not filters:
            rbg = alt(row - start)
            dcell(ws, row, 1, role["name"], bold=True, bg=rbg)
            dcell(ws, row, 2, "-", bg=rbg)
            dcell(ws, row, 3, "No filter defined", bg=rbg)
            ws.row_dimensions[row].height = 17
            row += 1
        else:
            for f in filters:
                rbg = alt(row - start)
                dcell(ws, row, 1, role["name"], bold=True, bg=rbg)
                dcell(ws, row, 2, f.get("table", ""), bg=rbg)
                dcell(ws, row, 3, f.get("expression", ""), bg=rbg, wrap=True)
                ws.row_dimensions[row].height = 17
                row += 1


def main():
    parser = argparse.ArgumentParser(description="Convert quality-summary.json to Excel.")
    parser.add_argument("--summary", default="ci-tools/quality-summary.json")
    parser.add_argument("--output",  default="ci-tools/quality-report.xlsx")
    args = parser.parse_args()

    summary_path = Path(args.summary)
    if not summary_path.exists():
        raise FileNotFoundError(f"quality-summary.json not found: {summary_path}")

    print(f"Reading: {summary_path}")
    data = json.loads(summary_path.read_text(encoding="utf-8-sig"))

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print("Building sheets...")
    build_summary_sheet(wb, data)
    build_bpa_violations_sheet(wb, data)
    build_bpa_summary_sheet(wb, data)
    build_inspector_sheet(wb, data)
    build_model_overview_sheet(wb, data)
    build_columns_sheet(wb, data)
    build_measures_sheet(wb, data)
    build_relationships_sheet(wb, data)
    build_sources_sheet(wb, data)
    build_roles_sheet(wb, data)

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    print(f"\nExcel report saved to: {out_path}")
    print(f"Sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()